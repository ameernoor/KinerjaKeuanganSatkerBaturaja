import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import io
import numpy as np
from pathlib import Path
import os
import base64
from github import Github
from github import Auth
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# define month order map
MONTH_ORDER = {
    "JANUARI": 1, "FEBRUARI": 2, "PEBRUARI": 2, "MARET": 3, "APRIL": 4, "MEI": 5, "JUNI": 6,
    "JULI": 7, "AGUSTUS": 8, "SEPTEMBER": 9, "OKTOBER": 10, 
    "NOVEMBER": 11, "NOPEMBER": 11, "DESEMBER": 12
}
# Konfigurasi halaman
st.set_page_config(
    page_title="Dashboard IKPA KPPN Baturaja",
    page_icon="📊",
    layout="wide"
)

# Inisialisasi session state untuk menyimpan data dan aktivitas
if 'data_storage' not in st.session_state:
    st.session_state.data_storage = {}

if 'activity_log' not in st.session_state:
    st.session_state.activity_log = []  # Each entry: dict with timestamp, action, period, status

# Fungsi untuk memproses file Excel
def process_excel_file(uploaded_file, year):
    """
    Memproses file Excel IKPA sesuai struktur yang telah ditentukan
    """
    try:
        df_raw = pd.read_excel(uploaded_file, header=None)
        
        # 1️⃣ Ekstrak bulan dari baris ke-2 (index 1)
        month_text = str(df_raw.iloc[1, 0])
        month = month_text.split(":")[-1].strip() if ":" in month_text else "UNKNOWN"
        
        # 2️⃣ Ekstrak data (baris ke-5 dst)
        df_data = df_raw.iloc[4:].reset_index(drop=True)
        df_data.columns = range(len(df_data.columns))
        
        processed_rows = []
        i = 0
        while i < len(df_data):
            if i + 3 >= len(df_data):
                break
            
            nilai_row = df_data.iloc[i]
            bobot_row = df_data.iloc[i + 1]
            nilai_akhir_row = df_data.iloc[i + 2]
            nilai_aspek_row = df_data.iloc[i + 3]
            
            # Ekstrak kolom
            no = nilai_row[0]
            kode_kppn = str(nilai_row[1]).strip("'") if pd.notna(nilai_row[1]) else ""
            kode_ba = str(nilai_row[2]).strip("'") if pd.notna(nilai_row[2]) else ""
            kode_satker = str(nilai_row[3]).strip("'") if pd.notna(nilai_row[3]) else ""
            uraian_satker = nilai_row[4] if pd.notna(nilai_row[4]) else ""
            
            aspek_perencanaan = nilai_aspek_row[6] if pd.notna(nilai_aspek_row[6]) else 0
            aspek_pelaksanaan = nilai_aspek_row[8] if pd.notna(nilai_aspek_row[8]) else 0
            aspek_hasil = nilai_aspek_row[12] if pd.notna(nilai_aspek_row[12]) else 0
            
            revisi_dipa = nilai_row[6] if pd.notna(nilai_row[6]) else 0
            deviasi_hal3 = nilai_row[7] if pd.notna(nilai_row[7]) else 0
            penyerapan = nilai_row[8] if pd.notna(nilai_row[8]) else 0
            belanja_kontraktual = nilai_row[9] if pd.notna(nilai_row[9]) else 0
            penyelesaian_tagihan = nilai_row[10] if pd.notna(nilai_row[10]) else 0
            pengelolaan_up = nilai_row[11] if pd.notna(nilai_row[11]) else 0
            capaian_output = nilai_row[12] if pd.notna(nilai_row[12]) else 0
            
            nilai_total = nilai_row[13] if pd.notna(nilai_row[13]) else 0
            konversi_bobot = nilai_row[14] if pd.notna(nilai_row[14]) else 0
            dispensasi_spm = nilai_row[15] if pd.notna(nilai_row[15]) else 0
            nilai_akhir = nilai_row[16] if pd.notna(nilai_row[16]) else 0

            # Simpan bobot & nilai terbobot
            bobot_dict = {
                'Revisi DIPA': bobot_row[6], 'Deviasi Halaman III DIPA': bobot_row[7],
                'Penyerapan Anggaran': bobot_row[8], 'Belanja Kontraktual': bobot_row[9],
                'Penyelesaian Tagihan': bobot_row[10], 'Pengelolaan UP dan TUP': bobot_row[11],
                'Capaian Output': bobot_row[12]
            }
            nilai_terbobot_dict = {
                'Revisi DIPA': nilai_akhir_row[6], 'Deviasi Halaman III DIPA': nilai_akhir_row[7],
                'Penyerapan Anggaran': nilai_akhir_row[8], 'Belanja Kontraktual': nilai_akhir_row[9],
                'Penyelesaian Tagihan': nilai_akhir_row[10], 'Pengelolaan UP dan TUP': nilai_akhir_row[11],
                'Capaian Output': nilai_akhir_row[12]
            }

            row_data = {
                'No': no, 'Kode KPPN': kode_kppn, 'Kode BA': kode_ba, 'Kode Satker': kode_satker,
                'Uraian Satker': uraian_satker,
                'Kualitas Perencanaan Anggaran': aspek_perencanaan,
                'Kualitas Pelaksanaan Anggaran': aspek_pelaksanaan,
                'Kualitas Hasil Pelaksanaan Anggaran': aspek_hasil,
                'Revisi DIPA': revisi_dipa, 'Deviasi Halaman III DIPA': deviasi_hal3,
                'Penyerapan Anggaran': penyerapan, 'Belanja Kontraktual': belanja_kontraktual,
                'Penyelesaian Tagihan': penyelesaian_tagihan, 'Pengelolaan UP dan TUP': pengelolaan_up,
                'Capaian Output': capaian_output,
                'Nilai Total': nilai_total, 'Konversi Bobot': konversi_bobot,
                'Dispensasi SPM (Pengurang)': dispensasi_spm,
                'Nilai Akhir (Nilai Total/Konversi Bobot)': nilai_akhir,
                'Bulan': month, 'Tahun': year,
                'Bobot': bobot_dict, 'Nilai Terbobot': nilai_terbobot_dict
            }
            processed_rows.append(row_data)
            i += 4

        df_processed = pd.DataFrame(processed_rows)
        df_processed = df_processed.sort_values('Nilai Akhir (Nilai Total/Konversi Bobot)', ascending=False)
        df_processed['Peringkat'] = range(1, len(df_processed) + 1)

        # Apply reference short names (if available)
        df_processed = apply_reference_short_names(df_processed)
        df_processed = create_satker_column(df_processed)  # Use helper function
        df_processed['Source'] = 'Upload'

        return df_processed, month, year

    except Exception as e:
        st.error(f"Error memproses file: {str(e)}")
        return None, None, None

# Save any file (Excel/template) to your GitHub repo
def save_file_to_github(file_bytes, filename, folder="data"):
    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        st.stop()
        st.error("❌ Gagal mengakses GitHub: GITHUB_TOKEN atau GITHUB_REPO tidak ditemukan di secrets.")
        return

    g = Github(auth=Auth.Token(token))
    repo = g.get_repo(repo_name)
    path = f"{folder}/{filename}"

    try:
        contents = repo.get_contents(path)
        repo.update_file(contents.path, f"Update {filename}", file_bytes, contents.sha)
        st.success(f"✅ File {filename} diperbarui di GitHub.")
    except Exception:
        repo.create_file(path, f"Upload {filename}", file_bytes)
        st.success(f"✅ File {filename} diunggah ke GitHub.")


# Load all uploaded data from GitHub (run on startup)
def load_data_from_github():
    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        st.stop()
        st.error("❌ Gagal mengakses GitHub: GITHUB_TOKEN atau GITHUB_REPO tidak ditemukan di secrets.")
        return

    g = Github(auth=Auth.Token(token))
    repo = g.get_repo(repo_name)

    try:
        contents = repo.get_contents("data")
    except Exception:
        st.info("📁 Folder 'data' belum ada di repository GitHub.")
        return

    st.session_state.data_storage = {}

    for file in contents:
        if not file.name.endswith(".xlsx"):
            continue

        decoded = base64.b64decode(file.content)
        df = pd.read_excel(io.BytesIO(decoded))
        parts = file.name.replace("IKPA_", "").replace(".xlsx", "").split("_")
        if len(parts) != 2:
            continue
        month, year = parts

        # Apply reference short names first
        df = apply_reference_short_names(df)
        # Then create Satker column consistently
        df = create_satker_column(df)

        numeric_cols = [
            'Nilai Akhir (Nilai Total/Konversi Bobot)', 'Nilai Total', 'Konversi Bobot',
            'Revisi DIPA', 'Deviasi Halaman III DIPA', 'Penyerapan Anggaran',
            'Belanja Kontraktual', 'Penyelesaian Tagihan', 'Pengelolaan UP dan TUP', 'Capaian Output'
        ]
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        df['Bulan'] = df.get('Bulan', month)
        df['Tahun'] = df.get('Tahun', year)
        df['Source'] = 'GitHub'
        df['Period'] = f"{month} {year}"
        month_num = MONTH_ORDER.get(month.upper(), 0)
        df['Period_Sort'] = f"{int(year):04d}-{month_num:02d}"

        if 'Peringkat' not in df.columns and 'Nilai Akhir (Nilai Total/Konversi Bobot)' in df.columns:
            df = df.sort_values('Nilai Akhir (Nilai Total/Konversi Bobot)', ascending=False)
            df['Peringkat'] = range(1, len(df) + 1)
        
        st.session_state.data_storage[(str(month), str(year))] = df

    st.success(f"✅ {len(st.session_state.data_storage)} file berhasil dimuat dari GitHub.")

# Path ke file template (akan diatur di session state)
TEMPLATE_PATH = r"C:\Users\KEMENKEU\Desktop\INDIKATOR PELAKSANAAN ANGGARAN.xlsx"

# Fungsi untuk membaca template Excel yang sudah ada
def get_template_file():
    """
    Membaca file template Excel yang sudah ada
    """
    try:
        # Cek apakah file template ada di path default
        if Path(TEMPLATE_PATH).exists():
            with open(TEMPLATE_PATH, 'rb') as f:
                return f.read()
        else:
            # Jika tidak ada, gunakan template dari session state (jika di-upload admin)
            if 'template_file' in st.session_state:
                return st.session_state.template_file
            else:
                return None
    except Exception as e:
        st.error(f"Error membaca template: {str(e)}")
        return None

# Fungsi visualisasi podium/bintang
def create_ranking_chart(df, title, top=True, limit=10):
    """
    Membuat visualisasi ranking dengan bar chart horizontal yang menarik
    (Sekarang menggunakan kolom 'Satker' untuk label agar unik)
    """
    if top:
        df_sorted = df.nlargest(limit, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
        color_scale = 'Greens'
        emoji = '🏆'
    else:
        df_sorted = df.nsmallest(limit, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
        color_scale = 'Reds'
        emoji = '⚠️'
    
    fig = go.Figure()
    
    colors = px.colors.sequential.Greens if top else px.colors.sequential.Reds
    
    # use 'Satker' for y labels to keep them unique
    fig.add_trace(go.Bar(
        y=df_sorted['Satker'],
        x=df_sorted['Nilai Akhir (Nilai Total/Konversi Bobot)'],
        orientation='h',
        marker=dict(
            color=df_sorted['Nilai Akhir (Nilai Total/Konversi Bobot)'],
            colorscale=color_scale,
            showscale=False
        ),
        text=df_sorted['Nilai Akhir (Nilai Total/Konversi Bobot)'].round(2),
        textposition='outside',
        hovertemplate='<b>%{y}</b><br>Nilai: %{x:.2f}<extra></extra>'
    ))
    
    fig.update_layout(
        title=f"{emoji} {title}",
        xaxis_title="Nilai Akhir",
        yaxis_title="",
        height=max(400, limit * 40),
        yaxis={'categoryorder': 'total ascending' if not top else 'total descending'},
        showlegend=False
    )
    
    return fig

# ============================================================
# 🧩 Improved Problem Chart (with sorting, sliders, and filters)
# ============================================================
def create_problem_chart(df, column, threshold, title, comparison='less', y_min=None, y_max=None, show_yaxis=True):
    """
    Membuat visualisasi vertikal untuk satker dengan masalah.
    - Menampilkan satker secara urut berdasarkan nilai (ascending)
    - Menghapus nilai 0 khusus untuk kolom Pengelolaan UP dan TUP
    - Tidak menampilkan label X-axis
    - Opsi pengaturan rentang Y-axis dari slider
    """
    if comparison == 'less':
        df_filtered = df[df[column] < threshold]
    else:
        df_filtered = df[df[column] > threshold]

    # 🧹 Khusus Pengelolaan UP dan TUP: abaikan nilai 0
    if "UP" in column.upper() and "TUP" in column.upper():
        df_filtered = df_filtered[df_filtered[column] != 0]

    if len(df_filtered) == 0:
        return None

    # 🧮 Urutkan berdasarkan nilai
    df_filtered = df_filtered.sort_values(by=column, ascending=False)

    # Nilai min/max data
    min_val = df_filtered[column].min()
    max_val = df_filtered[column].max()

    # 🧩 Gunakan nilai dari slider jika ada
    if y_min is None:
        y_min = max(0, int(min_val) - 5)
    if y_max is None:
        y_max = min(110, int(max_val) + 5)

    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=df_filtered['Satker'],
        y=df_filtered[column],
        marker=dict(
            color=df_filtered[column],
            colorscale='OrRd_r',
            showscale=True,
            cmin=min_val,
            cmax=max_val,
        ),
        text=df_filtered[column].round(2),
        textposition='outside',
        hovertemplate='<b>%{x}</b><br>Nilai: %{y:.2f}<extra></extra>'
    ))

    # 🔹 Garis target threshold
    fig.add_hline(
        y=threshold,
        line_dash="dash",
        line_color="red",
        annotation_text=f"Target: {threshold}",
        annotation_position="top right"
    )

    # Layout
    fig.update_layout(
        title=f"⚠️ {title}",
        xaxis_title="",  # 🚫 Hapus label X-axis
        yaxis_title="Nilai" if show_yaxis else "",
        yaxis_range=[y_min, y_max],
        xaxis_tickangle=90,
        height=500,
        margin=dict(l=10, r=10, t=50, b=120),
        showlegend=False,
    )

    if not show_yaxis:
        fig.update_yaxes(showticklabels=False)

    return fig

# ===============================================
# 🧩 Helper to apply reference short names (Fixed)
# ===============================================
def apply_reference_short_names(df):
    """
    Apply reference short names to dataframe with advanced error tracking.
    If errors found, triggers Excel download with error details.
    
    This function only adds 'Uraian Satker-SINGKAT' and 'Uraian Satker Final' columns.
    The 'Satker' column should be created by the caller for consistency.
    """
    if 'reference_df' not in st.session_state or st.session_state.reference_df is None:
        # If no reference data, create 'Uraian Satker Final' from existing 'Uraian Satker'
        if 'Uraian Satker Final' not in df.columns:
            df['Uraian Satker Final'] = df.get('Uraian Satker', '')
        return df

    ref = st.session_state.reference_df.copy()

    # 🩹 Ensure same dtype before merging
    df['Kode Satker'] = df['Kode Satker'].astype(str).str.strip()
    ref['Kode Satker'] = ref['Kode Satker'].astype(str).str.strip()

    try:
        # Merge with indicator to track unmatched records
        df_merged = df.merge(
            ref[['Kode Satker', 'Uraian Satker-SINGKAT']], 
            on='Kode Satker', 
            how='left',
            indicator=True
        )
        
        # Identify rows without matching reference
        missing_refs = df_merged[df_merged['_merge'] == 'left_only'].copy()
        
        # If there are missing references, prepare error report
        if len(missing_refs) > 0:
            # Create error dataframe with required columns
            error_cols = ['Tahun', 'Bulan', 'Kode Satker', 'Uraian Satker']
            error_df = missing_refs[error_cols].drop_duplicates().sort_values(['Tahun', 'Bulan', 'Kode Satker'])
            
            # Add row numbers for reference
            error_df.insert(0, 'No', range(1, len(error_df) + 1))
            
            # Create Excel file in memory using openpyxl
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                error_df.to_excel(writer, sheet_name='Satker Tidak Ditemukan', index=False)
                
                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Satker Tidak Ditemukan']
                
                # Define styles
                header_fill = PatternFill(start_color='D7E4BD', end_color='D7E4BD', fill_type='solid')
                header_font = Font(bold=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                center_align = Alignment(horizontal='center', vertical='center')
                left_align = Alignment(horizontal='left', vertical='center')
                
                # Apply header formatting
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = center_align
                
                # Apply cell formatting and borders
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border
                        cell.alignment = left_align
                
                # Adjust column widths
                worksheet.column_dimensions['A'].width = 8   # No
                worksheet.column_dimensions['B'].width = 12  # Tahun
                worksheet.column_dimensions['C'].width = 15  # Bulan
                worksheet.column_dimensions['D'].width = 20  # Kode Satker
                worksheet.column_dimensions['E'].width = 50  # Uraian Satker
            
            excel_data = output.getvalue()
            
            # Display single warning message with download button
            st.warning(
                "⚠️ **Referensi singkatan untuk satker sebagaimana daftar terlampir tidak ditemukan.** "
                f"\n\n📊 Total {len(error_df)} satker tidak ditemukan dalam database referensi. "
                "\n\nMohon agar admin memeriksa dan mengupdate database referensi."
            )
            
            # Provide download button
            st.download_button(
                label="📥 Download Daftar Satker Tidak Ditemukan (Excel)",
                data=excel_data,
                file_name=f"satker_tidak_ditemukan_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        
        # Apply short names - create 'Uraian Satker Final' column
        df_merged['Uraian Satker Final'] = df_merged['Uraian Satker-SINGKAT'].fillna(
            df_merged.get('Uraian Satker', '')
        )
        
        # Remove merge indicator column
        df_merged = df_merged.drop(columns=['_merge'])
        
        return df_merged
        
    except Exception as e:
        st.error(f"❌ Gagal menerapkan nama singkat: {e}")
        # Fallback: create 'Uraian Satker Final' from existing 'Uraian Satker'
        if 'Uraian Satker Final' not in df.columns:
            df['Uraian Satker Final'] = df.get('Uraian Satker', '')
        return df


# ===============================================
# 📝 UPDATED: Helper function to create Satker column consistently
# ===============================================
def create_satker_column(df):
    """
    Creates 'Satker' column consistently across all data sources.
    Should be called after apply_reference_short_names().
    """
    if 'Uraian Satker Final' not in df.columns:
        df['Uraian Satker Final'] = df.get('Uraian Satker', '')
    
    df['Satker'] = (
        df['Uraian Satker Final'].astype(str) + 
        ' (' + df['Kode Satker'].astype(str) + ')'
    )
    return df

# HALAMAN 1: DASHBOARD UTAMA
def page_dashboard():
    st.title("📊 Dashboard Utama IKPA Satker Mitra KPPN Baturaja")
    
    if not st.session_state.data_storage:
        st.warning("⚠️ Belum ada data yang diunggah. Silakan unggah data melalui halaman Admin.")
        return
    
    # Dapatkan data terbaru
    all_periods = sorted(
        st.session_state.data_storage.keys(),
        key=lambda x: (int(x[1]), MONTH_ORDER.get(x[0].upper(), 0)),
        reverse=True
    )
    
    if not all_periods:
        st.warning("⚠️ Belum ada data yang tersedia.")
        return

    # Pembatas & Judul
    st.markdown("---")
    st.markdown("## 🎯 Highlights Kinerja Satker")

    # ===============================
    # 🎯 Periode & Ringkasan Metrik — Single Row Layout
    # ===============================
    col_period, col1, col2, col3, col4 = st.columns([1, 1, 1, 1, 1])

    # 📅 Period selection (takes more width)
    with col_period:
        selected_period = st.selectbox(
            "Pilih Periode",
            options=all_periods,
            index=0,
            format_func=lambda x: f"{x[0].capitalize()} {x[1]}"
        )

    df = st.session_state.data_storage[selected_period]

    # 📊 Compact summary metrics
    with col1:
        st.metric("📋 Total Satker", len(df))
    with col2:
        avg_score = df['Nilai Akhir (Nilai Total/Konversi Bobot)'].mean()
        st.metric("📈 Rata-rata Nilai", f"{avg_score:.2f}")
    with col3:
        perfect_count = len(df[df['Nilai Akhir (Nilai Total/Konversi Bobot)'] == 100])
        st.metric("⭐ Nilai 100", perfect_count)
    with col4:
        below_89 = len(df[df['Nilai Akhir (Nilai Total/Konversi Bobot)'] < 89])
        st.metric("⚠️ Nilai < 89 (Predikat Belum Baik)", below_89)
    
    # ===============================
    # 📊 Ranking Charts — Compact Horizontal Strip (4 in 1)
    # ===============================

    # 🎚️ User control for Y-axis range
    st.markdown("###### Atur Skala Nilai (Sumbu Y)")

    col_min, col_max = st.columns(2)
    with col_min:
        y_min = st.slider(
            "Nilai Minimum (Y-Axis)",
            min_value=0,
            max_value=50,
            value=50,  # Default to start from 50
            step=1,
        )
    with col_max:
        y_max = st.slider(
            "Nilai Maksimum (Y-Axis)",
            min_value=51,
            max_value=110,
            value=110,
            step=1,
        )

    # ===============================
    # Data preparation
    # ===============================
    df_with_kontrak = df[df['Belanja Kontraktual'] != 0]
    df_without_kontrak = df[df['Belanja Kontraktual'] == 0]

    # ===============================
    # Chart creator helper
    # ===============================
    def make_column_chart(data, title, color_scale, y_min, y_max, limit=10, show_yaxis=False):
        """Creates a compact vertical bar (column) chart with consistent style"""
        fig = px.bar(
            data.nlargest(limit, 'Nilai Akhir (Nilai Total/Konversi Bobot)'),
            x='Satker',
            y='Nilai Akhir (Nilai Total/Konversi Bobot)',
            color='Nilai Akhir (Nilai Total/Konversi Bobot)',
            color_continuous_scale=color_scale,
            title=title,
        )
        fig.update_layout(
            yaxis_range=[y_min, y_max],
            xaxis_tickangle=90,  # 🔹 Rotate labels vertically
            height=500,
            margin=dict(l=10, r=10, t=40, b=80),
            coloraxis_showscale=False,
            showlegend=False,
        )

        # 🔹 Optional Y-axis display
        if show_yaxis:
            fig.update_yaxes(title_text="Nilai IKPA", showticklabels=True)
        else:
            fig.update_yaxes(title_text="", showticklabels=False)

        # 🔹 Remove "Satker" label from X-axis
        fig.update_xaxes(title_text="")

        fig.update_traces(
            texttemplate='%{y:.1f}',
            textposition='outside',
            hovertemplate='<b>%{x}</b><br>Nilai: %{y:.2f}<extra></extra>'
        )

        return fig

    # ===============================
    # 4 charts side by side (4x1 layout)
    # ===============================
    col1, col2, col3, col4 = st.columns(4)

    # 1️⃣ Top 10 with kontraktual
    with col1:
        st.markdown("##### 🏆 10 Satker Terbaik (Dengan Kontraktual)")
        if len(df_with_kontrak) > 0:
            top_with = df_with_kontrak.nlargest(10, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
            fig1 = make_column_chart(top_with, "", "greens", y_min, y_max, show_yaxis=True)
            st.plotly_chart(fig1, use_container_width=True)
        else:
            st.info("Tidak ada data.")

    # 2️⃣ Top 10 without kontraktual
    with col2:
        st.markdown("##### 🏆 10 Satker Terbaik (Tanpa Kontraktual)")
        if len(df_without_kontrak) > 0:
            top_without = df_without_kontrak.nlargest(10, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
            fig2 = make_column_chart(top_without, "", "greens", y_min, y_max, show_yaxis=False)
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Tidak ada data.")

    # 3️⃣ Bottom 10 with kontraktual
    with col3:
        st.markdown("##### 📉 10 Satker Terendah (Dengan Kontraktual)")
        if len(df_with_kontrak) > 0:
            bottom_with = df_with_kontrak.nsmallest(10, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
            fig3 = make_column_chart(bottom_with, "", "orrd_r", y_min, y_max, show_yaxis=False)
            st.plotly_chart(fig3, use_container_width=True)
        else:
            st.info("Tidak ada data.")

    # 4️⃣ Bottom 10 without kontraktual
    with col4:
        st.markdown("##### 📉 10 Satker Terendah (Tanpa Kontraktual)")
        if len(df_without_kontrak) > 0:
            bottom_without = df_without_kontrak.nsmallest(10, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
            fig4 = make_column_chart(bottom_without, "", "orrd_r", y_min, y_max, show_yaxis=False)
            st.plotly_chart(fig4, use_container_width=True)
        else:
            st.info("Tidak ada data.")

    st.markdown("---")
    
    # Satker dengan masalah
    st.subheader("🚨 Satker yang Memerlukan Perhatian Khusus")

    # 🎚️ Pengaturan Sumbu Y
    st.markdown("###### Atur Skala Nilai (Sumbu Y)")
    col_min, col_max = st.columns(2)
    with col_min:
        y_min_dev = st.slider(
            "Nilai Minimum (Y-Axis)",
            min_value=0,
            max_value=50,
            value=40,
            step=1,
            key="ymin_dev"
        )
    with col_max:
        y_max_dev = st.slider(
            "Nilai Maksimum (Y-Axis)",
            min_value=51,
            max_value=110,
            value=110,
            step=1,
            key="ymax_dev"
        )

    # Deviasi Hal 3 DIPA
    fig_dev = create_problem_chart(
        df, 
        'Deviasi Halaman III DIPA', 
        90, 
        "Deviasi Hal 3 DIPA Belum Optimal (< 90)",
        'less',
        y_min=y_min_dev,
        y_max=y_max_dev,
        show_yaxis=True
    )
    if fig_dev:
        st.plotly_chart(fig_dev, use_container_width=True)
    else:
        st.success("✅ Semua satker sudah optimal untuk Deviasi Hal 3 DIPA")
    
    st.markdown("---")
    
    # Tabel detail
    st.subheader("📋 Tabel Detail Satker")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        view_mode = st.radio(
            "Tampilan",
            options=['aspek', 'komponen'],
            format_func=lambda x: 'Berdasarkan Aspek' if x == 'aspek' else 'Berdasarkan Komponen',
            horizontal=True
        )
    
    with col2:
        if view_mode == 'komponen':
            value_type = st.selectbox(
                "Jenis Nilai",
                options=['Nilai', 'Bobot', 'Nilai Terbobot']
            )
    
    # Kolom yang ditampilkan
    display_columns = ['Peringkat', 'Kode BA', 'Kode Satker', 'Uraian Satker']
    
    if view_mode == 'aspek':
        display_columns += [
            'Kualitas Perencanaan Anggaran',
            'Kualitas Pelaksanaan Anggaran',
            'Kualitas Hasil Pelaksanaan Anggaran'
        ]
        df_display = df[display_columns + ['Nilai Total', 'Konversi Bobot', 
                                            'Dispensasi SPM (Pengurang)', 
                                            'Nilai Akhir (Nilai Total/Konversi Bobot)']].copy()
    else:
        component_cols = [
            'Revisi DIPA', 'Deviasi Halaman III DIPA', 'Penyerapan Anggaran',
            'Belanja Kontraktual', 'Penyelesaian Tagihan', 
            'Pengelolaan UP dan TUP', 'Capaian Output'
        ]
        
        df_display = df[display_columns + ['Nilai Total', 'Konversi Bobot', 
                                            'Dispensasi SPM (Pengurang)', 
                                            'Nilai Akhir (Nilai Total/Konversi Bobot)']].copy()
        
        if value_type == 'Nilai':
            for col in component_cols:
                df_display[col] = df[col]
        elif value_type == 'Bobot':
            for col in component_cols:
                df_display[col] = df['Bobot'].apply(lambda x: x.get(col, 0) if isinstance(x, dict) else 0)
        else:  # Nilai Terbobot
            for col in component_cols:
                df_display[col] = df['Nilai Terbobot'].apply(lambda x: x.get(col, 0) if isinstance(x, dict) else 0)
        
        # Reorder kolom
        final_cols = display_columns + component_cols + ['Nilai Total', 'Konversi Bobot', 
                                                          'Dispensasi SPM (Pengurang)', 
                                                          'Nilai Akhir (Nilai Total/Konversi Bobot)']
        df_display = df_display[final_cols]
    
    # Styling untuk tabel
    def highlight_top(s):
        if s.name == 'Peringkat':
            return ['background-color: gold' if v <= 3 else '' for v in s]
        return ['' for _ in s]
    
    st.dataframe(
        df_display.style.apply(highlight_top).format(precision=2),
        use_container_width=True,
        height=600
    )

# HALAMAN 2: DASHBOARD INTERNAL KPPN (Protected)
def page_trend():
    st.title("🏛️ Early Warning System Kinerja Keuangan Satker")

    # 🔒 Access restriction (same password as Admin page)
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.warning("🔒 Halaman ini memerlukan autentikasi Admin untuk diakses.")
        password = st.text_input("Masukkan Password", type="password")
        if st.button("Login"):
            if password == "109KPPN":
                st.session_state.authenticated = True
                st.success("✅ Login berhasil! Silakan akses halaman ini.")
                st.rerun()
            else:
                st.error("❌ Password salah!")
        return
    
    if not st.session_state.data_storage:
        st.warning("⚠️ Belum ada data yang diunggah. Silakan unggah data melalui halaman Admin.")
        return
    
    # Gabungkan semua data
    all_data = []
    for period, df in st.session_state.data_storage.items():
        df_copy = df.copy()
        # ensure Period & Period_Sort exist
        df_copy['Period'] = f"{period[0]} {period[1]}"
        df_copy['Period_Sort'] = f"{period[1]}-{period[0]}"
        all_data.append(df_copy)
    
    if not all_data:
        st.warning("⚠️ Belum ada data historis yang tersedia.")
        return
    
    df_all = pd.concat(all_data, ignore_index=True)
      
    # Analisis tren dan Early Warning System
    # Gunakan data periode terkini
    latest_period = sorted(st.session_state.data_storage.keys(), key=lambda x: (int(x[1]), MONTH_ORDER.get(x[0].upper(), 0)), reverse=True)[0]
    df_latest = st.session_state.data_storage[latest_period]

    st.markdown("---")
    st.subheader("🚨 Satker yang Memerlukan Perhatian Khusus")

    # 🎚️ Pengaturan Sumbu Y
    st.markdown("###### Atur Skala Nilai (Sumbu Y)")
    col_min, col_max = st.columns(2)
    with col_min:
        y_min_int = st.slider(
            "Nilai Minimum (Y-Axis)",
            min_value=0,
            max_value=50,
            value=50,
            step=1,
            key="ymin_internal"
        )
    with col_max:
        y_max_int = st.slider(
            "Nilai Maksimum (Y-Axis)",
            min_value=51,
            max_value=110,
            value=110,
            step=1,
            key="ymax_internal"
        )

    # 📊 Highlights Kinerja Satker yang Perlu Perhatian Khusus
    col1, col2 = st.columns(2)

    with col1:
        fig_up = create_problem_chart(
            df_latest,
            'Pengelolaan UP dan TUP',
            100,
            "Pengelolaan UP dan TUP Belum Optimal (< 100)",
            'less',
            y_min=y_min_int,
            y_max=y_max_int,
            show_yaxis=True  # Left chart shows Y-axis
        )
        if fig_up:
            st.plotly_chart(fig_up, use_container_width=True)
        else:
            st.success("✅ Semua satker sudah optimal untuk Pengelolaan UP dan TUP")

    with col2:
        fig_output = create_problem_chart(
            df_latest,
            'Capaian Output',
            100,
            "Capaian Output Belum Optimal (< 100)",
            'less',
            y_min=y_min_int,
            y_max=y_max_int,
            show_yaxis=False  # Right chart hides Y-axis
        )
        if fig_output:
            st.plotly_chart(fig_output, use_container_width=True)
        else:
            st.success("✅ Semua satker sudah optimal untuk Capaian Output")
    
    warnings = []

    st.markdown("---")
# Analisis Tren
    st.subheader("📈 Analisis Tren")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # 🔍 DETAILED ERROR CHECKING
        st.write("🔍 Checking data quality...")
        
        # Map month names to numbers
        df_all['Month_Num'] = df_all['Bulan'].str.strip().str.upper().map(MONTH_ORDER)
        
        # Check for unmapped months
        missing_months = df_all[df_all['Month_Num'].isna()]
        if len(missing_months) > 0:
            st.error("❌ **DITEMUKAN BULAN YANG TIDAK VALID:**")
            
            # Group by period to show which files have issues
            problem_periods = missing_months.groupby(['Bulan', 'Tahun']).size().reset_index(name='Count')
            
            for _, row in problem_periods.iterrows():
                st.warning(f"⚠️ Periode **{row['Bulan']} {row['Tahun']}** - Nama bulan '{row['Bulan']}' tidak dikenali (ditemukan di {row['Count']} baris)")
            
            st.info("""
            **Solusi:**
            1. Periksa file Excel untuk periode yang bermasalah
            2. Pastikan nama bulan sesuai format: JANUARI, FEBRUARI, MARET, dst (huruf besar)
            3. Upload ulang file yang bermasalah dari halaman Admin
            """)
            
            # Show expected month names
            with st.expander("📋 Lihat format bulan yang valid"):
                st.write("Format yang diterima:")
                st.code(", ".join(MONTH_ORDER.keys()))
            
            # Option to proceed with cleaned data
            if st.checkbox("⚠️ Abaikan data bermasalah dan lanjutkan"):
                df_all = df_all.dropna(subset=['Month_Num'])
                st.info(f"✅ Data dibersihkan. Sisa {len(df_all)} baris.")
            else:
                st.stop()
        
        # Check for invalid years
        invalid_years = df_all[df_all['Tahun'].isna()]
        if len(invalid_years) > 0:
            st.error("❌ **DITEMUKAN TAHUN YANG TIDAK VALID:**")
            
            problem_periods = invalid_years.groupby(['Bulan']).size().reset_index(name='Count')
            for _, row in problem_periods.iterrows():
                st.warning(f"⚠️ Bulan **{row['Bulan']}** - Tahun tidak valid (ditemukan di {row['Count']} baris)")
            
            st.stop()
        
        # Try to create Period_Sort with detailed error handling
        try:
            # Convert to int safely
            df_all['Tahun_Int'] = df_all['Tahun'].astype(int)
            df_all['Month_Num_Int'] = df_all['Month_Num'].astype(int)
            
            # Create Period_Sort
            df_all['Period_Sort'] = df_all.apply(
                lambda x: f"{x['Tahun_Int']:04d}-{x['Month_Num_Int']:02d}", 
                axis=1
            )
            
            st.success(f"✅ Data valid - {len(df_all)} baris dari {df_all['Period'].nunique()} periode")
            
        except Exception as e:
            st.error(f"❌ **ERROR saat membuat Period_Sort:** {str(e)}")
            
            # Show problematic rows
            st.write("**Baris yang bermasalah:**")
            problem_cols = ['Bulan', 'Tahun', 'Month_Num', 'Kode Satker', 'Uraian Satker']
            st.dataframe(df_all[problem_cols].head(20))
            
            st.stop()
        
        # Now create the selectbox
        available_periods = sorted(df_all['Period_Sort'].unique())
        start_period = st.selectbox(
            "Periode Awal",
            options=available_periods,
            index=0
        )
    
    with col2:
        end_period = st.selectbox(
            "Periode Akhir",
            options=available_periods,
            index=len(available_periods) - 1
        )
    
    # Filter berdasarkan periode
    df_filtered = df_all[
        (df_all['Period_Sort'] >= start_period) & 
        (df_all['Period_Sort'] <= end_period)
    ]
    
    with col3:
        # Pilihan metrik
        metric_options = {
            'Nilai Akhir (Nilai Total/Konversi Bobot)': 'Nilai Akhir (Nilai Total/Konversi Bobot)',
            'Kualitas Perencanaan Anggaran': 'Kualitas Perencanaan Anggaran',
            'Kualitas Pelaksanaan Anggaran': 'Kualitas Pelaksanaan Anggaran',
            'Kualitas Hasil Pelaksanaan Anggaran': 'Kualitas Hasil Pelaksanaan Anggaran',
            'Revisi DIPA': 'Revisi DIPA',
            'Deviasi Halaman III DIPA': 'Deviasi Halaman III DIPA',
            'Penyerapan Anggaran': 'Penyerapan Anggaran',
            'Belanja Kontraktual': 'Belanja Kontraktual',
            'Penyelesaian Tagihan': 'Penyelesaian Tagihan',
            'Pengelolaan UP dan TUP': 'Pengelolaan UP dan TUP',
            'Capaian Output': 'Capaian Output'
        }
        
        selected_metric = st.selectbox(
            "Metrik yang Ditampilkan",
            options=list(metric_options.keys()),
            index=0
        )
    
    # Pilih satker
    # All keys are (month_str, year_str). To sort by year then month, create sortable key:
    def period_sort_key(k):
        mon, yr = k
        # convert year to int if possible, month remain string but sorting will be stable for same year
        try:
            y = int(yr)
        except Exception as e:
            st.warning(f"⚠️ Tidak bisa convert tahun '{yr}' untuk periode {mon}: {e}")
            y = 0
        return (y, mon)

    try:
        latest_period = sorted(st.session_state.data_storage.keys(), key=period_sort_key, reverse=True)[0]
        latest_df = st.session_state.data_storage[latest_period].copy()
    except Exception as e:
        st.error(f"❌ Error mendapatkan periode terbaru: {e}")
        st.write("**Periode yang tersedia:**")
        st.write(list(st.session_state.data_storage.keys()))
        st.stop()
    
    # Make sure 'Kode Satker' exists and is a string
    if 'Kode Satker' in latest_df.columns:
        latest_df['Kode Satker'] = latest_df['Kode Satker'].astype(str)
    else:
        latest_df['Kode Satker'] = latest_df.index.astype(str)

    bottom_10_default = latest_df.nsmallest(10, 'Nilai Akhir (Nilai Total/Konversi Bobot)')['Kode Satker'].astype(str).tolist()
    
    # use the new 'Satker' column for selection (unique)
    all_satker = sorted(df_all['Satker'].unique())
    selected_satker = st.multiselect(
        "Pilih Satker",
        options=all_satker,
        default=[s for s in all_satker if any(str(code) in s for code in bottom_10_default)][:10]
    )
    
    if not selected_satker:
        st.warning("Silakan pilih minimal satu satker untuk melihat tren.")
        return
    
    # Filter berdasarkan satker (use 'Satker' to avoid duplicate names)
    df_plot = df_filtered[df_filtered['Satker'].isin(selected_satker)]
    
    # Buat line chart
    fig = go.Figure()
    
    try:
        for satker in selected_satker:
            df_satker = df_plot[df_plot['Satker'] == satker].sort_values('Period_Sort')

            # Ensure x-axis uses correct chronological month order
            categories = [f"{m} {y}" for y, m in sorted(
                {(int(x['Tahun']), x['Bulan'].upper()) for _, x in df_all.iterrows()},
                key=lambda t: (t[0], MONTH_ORDER.get(t[1], 0))
            )]
            
            fig.add_trace(go.Scatter(
                x=pd.Categorical(
                    df_satker['Period'],
                    categories=categories,
                    ordered=True
                ),
                y=df_satker[selected_metric],
                mode='lines+markers',
                name=satker,
                hovertemplate='<b>%{fullData.name}</b><br>Periode: %{x}<br>Nilai: %{y:.2f}<extra></extra>'
            ))
    except Exception as e:
        st.error(f"❌ Error membuat chart: {str(e)}")
        st.write("**Debug Info:**")
        st.write(f"Selected satker: {selected_satker}")
        st.write(f"df_plot shape: {df_plot.shape}")
        st.write(f"Unique periods in df_plot: {df_plot['Period'].unique()}")
        st.stop()
    
    fig.update_layout(
        title=f"Tren {selected_metric}",
        xaxis_title="Periode",
        yaxis_title="Nilai",
        height=600,
        hovermode='x unified',
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1,
            xanchor="left",
            x=1.02
        )
    )
    
    st.plotly_chart(fig, use_container_width=True)

    # Early Warning Satker Tren Menurun
    warnings = []  # Initialize warnings list
    
    for satker in selected_satker:
        df_satker = df_plot[df_plot['Satker'] == satker].sort_values('Period_Sort')
        
        if len(df_satker) >= 2:
            values = df_satker[selected_metric].values
            
            # Cek tren menurun (2 periode terakhir)
            if len(values) >= 2:
                last_value = values[-1]
                prev_value = values[-2]
                
                if last_value < prev_value:
                    decrease = prev_value - last_value
                    warnings.append({
                        'Satker': satker,
                        'Metrik': selected_metric,
                        'Nilai Sebelumnya': prev_value,
                        'Nilai Terkini': last_value,
                        'Penurunan': decrease
                    })
    
    if warnings:
        st.warning(f"⚠️ Ditemukan {len(warnings)} satker dengan tren menurun!")
        
        for w in warnings:
            st.markdown(f"""
            **{w['Satker']}**  
            - Metrik: {w['Metrik']}
            - Nilai sebelumnya: {w['Nilai Sebelumnya']:.2f}
            - Nilai terkini: {w['Nilai Terkini']:.2f}
            - Penurunan: {w['Penurunan']:.2f} poin
            """)
            st.markdown("---")
    else:
        st.success("✅ Tidak ada satker dengan tren menurun pada periode yang dipilih!")
        
# ============================================================
# 🔐 HALAMAN 3: ADMIN (Revised with integrated Reference Upload)
# ============================================================
def page_admin():
    st.title("🔐 Halaman Administrasi")
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.warning("🔒 Halaman ini memerlukan autentikasi")
        password = st.text_input("Masukkan Password", type="password")
        if st.button("Login"):
            if password == "109KPPN":
                st.session_state.authenticated = True
                st.success("✅ Login berhasil!")
                st.rerun()
            else:
                st.error("❌ Password salah!")
        return

    st.success("✅ Anda telah login sebagai Admin")

    # 🧩 Debug GitHub connection
    with st.expander("🧩 Debug GitHub Connection"):
        try:
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]
            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)
            st.success(f"Terhubung ke GitHub repo: {repo.full_name}")
        except Exception as e:
            st.error(f"❌ Gagal terhubung ke GitHub: {e}")

    if st.button("🚪 Logout"):
        st.session_state.authenticated = False
        st.rerun()

    st.markdown("---")
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📤 Upload Data",
        "🗑️ Hapus Data",
        "📥 Download Data",
        "📋 Download Template",
        "🕓 Riwayat Aktivitas"
    ])

    # ============================================================
    # TAB 1: UPLOAD DATA (including Reference Upload)
    # ============================================================
    with tab1:
        # Submenu Upload Data Bulanan
        st.subheader("📤 Upload Data Bulanan IKPA")

        upload_year = st.selectbox(
            "Pilih Tahun",
            list(range(2020, 2031)),
            index=list(range(2020, 2031)).index(datetime.now().year)
        )

        uploaded_file = st.file_uploader("Pilih file Excel IKPA", type=['xlsx', 'xls'])

        if uploaded_file is not None:
            try:
                df_temp = pd.read_excel(uploaded_file, header=None)
                month_text = str(df_temp.iloc[1, 0])
                month_preview = month_text.split(":")[-1].strip() if ":" in month_text else "UNKNOWN"
                period_key_preview = (str(month_preview), str(upload_year))
                uploaded_file.seek(0)

                if period_key_preview in st.session_state.data_storage:
                    st.warning(f"⚠️ Data untuk **{month_preview} {upload_year}** sudah ada.")
                    confirm_replace = st.checkbox(
                        "✅ Ganti data yang sudah ada.",
                        key=f"confirm_replace_{month_preview}_{upload_year}"
                    )
                else:
                    confirm_replace = True
                    st.info(f"📝 Akan mengunggah data baru untuk periode: **{month_preview} {upload_year}**")

            except Exception as e:
                st.error(f"❌ Gagal membaca preview file: {e}")
                confirm_replace = False

            if st.button("🔄 Proses Data IKPA", type="primary", disabled=not confirm_replace):
                with st.spinner("Memproses data..."):
                    df_processed, month, year = process_excel_file(uploaded_file, upload_year)
                    if df_processed is None:
                        st.error("❌ Gagal memproses file.")
                        st.stop()

                    period_key = (str(month), str(year))
                    filename = f"IKPA_{month}_{year}.xlsx"

                    try:
                        df_processed['Kode Satker'] = df_processed['Kode Satker'].astype(str)
                        st.session_state.data_storage[period_key] = df_processed

                        excel_bytes = io.BytesIO()
                        with pd.ExcelWriter(excel_bytes, engine='openpyxl') as writer:
                            df_excel = df_processed.drop(['Bobot', 'Nilai Terbobot'], axis=1, errors='ignore')
                            df_excel.to_excel(writer, index=False, sheet_name='Data IKPA')
                        excel_bytes.seek(0)

                        save_file_to_github(excel_bytes.getvalue(), filename, folder="data")

                        st.success(f"✅ Data {month} {year} berhasil disimpan.")
                        st.snow()

                        st.session_state.activity_log.append({
                            "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "Aksi": "Upload",
                            "Periode": f"{month} {year}",
                            "Status": "✅ Sukses"
                        })
                    except Exception as e:
                        st.error(f"❌ Gagal menyimpan ke GitHub: {e}")


        # Sub Menu Upload Data Referensi
        st.markdown("---")
        st.subheader("📚 Upload / Perbarui Data Referensi Satker & K/L")
        st.info("""
        - File referensi ini berisi kolom: **Kode BA, K/L, Kode Satker, Uraian Satker-SINGKAT, Uraian Satker-LENGKAP**  
        - Saat diupload, sistem akan **menggabungkan** dengan data lama:  
          🔹 Jika `Kode Satker` sudah ada → baris lama akan **diganti**  
          🔹 Jika `Kode Satker` belum ada → akan **ditambahkan baru**
        """)

        uploaded_ref = st.file_uploader(
            "📤 Pilih File Data Referensi Satker & K/L",
            type=['xlsx', 'xls'],
            key="ref_upload"
        )

        if uploaded_ref is not None:
            try:
                new_ref = pd.read_excel(uploaded_ref)
                new_ref.columns = [c.strip() for c in new_ref.columns]

                required = ['Kode BA', 'K/L', 'Kode Satker', 'Uraian Satker-SINGKAT', 'Uraian Satker-LENGKAP']
                if not all(col in new_ref.columns for col in required):
                    st.error("❌ Kolom wajib tidak lengkap dalam file referensi.")
                    st.stop()

                new_ref['Kode Satker'] = new_ref['Kode Satker'].astype(str)

                # Gabungkan atau buat baru
                if 'reference_df' in st.session_state:
                    old_ref = st.session_state.reference_df.copy()
                    merged = pd.concat([old_ref, new_ref]).drop_duplicates(subset=['Kode Satker'], keep='last')
                    st.session_state.reference_df = merged
                    st.success(f"✅ Data Referensi diperbarui ({len(merged)} total baris).")
                else:
                    st.session_state.reference_df = new_ref
                    st.success(f"✅ Data Referensi baru dimuat ({len(new_ref)} baris).")

                st.dataframe(st.session_state.reference_df.tail(10), use_container_width=True)

                # 🧩 Save merged reference data permanently to GitHub
                try:
                    excel_bytes_ref = io.BytesIO()
                    with pd.ExcelWriter(excel_bytes_ref, engine='openpyxl') as writer:
                        st.session_state.reference_df.to_excel(writer, index=False, sheet_name='Data Referensi')
                    excel_bytes_ref.seek(0)

                    save_file_to_github(
                        excel_bytes_ref.getvalue(),
                        "Template_Data_Referensi.xlsx",
                        folder="templates"
                    )
                    st.success("💾 Data Referensi berhasil disimpan ke GitHub (templates/Template_Data_Referensi.xlsx).")
                except Exception as e:
                    st.error(f"❌ Gagal menyimpan Data Referensi ke GitHub: {e}")

            except Exception as e:
                st.error(f"❌ Gagal memproses Data Referensi: {e}")


    # ============================================================
    # TAB 2: HAPUS DATA
    # ============================================================
    with tab2:
        st.subheader("🗑️ Hapus Data Bulanan")
        if not st.session_state.data_storage:
            st.info("ℹ️ Belum ada data tersimpan.")
        else:
            available_periods = sorted(st.session_state.data_storage.keys(), reverse=True)
            period_to_delete = st.selectbox(
                "Pilih periode yang akan dihapus",
                options=available_periods,
                format_func=lambda x: f"{x[0].capitalize()} {x[1]}"
            )
            month, year = period_to_delete
            filename = f"data/IKPA_{month}_{year}.xlsx"

            confirm_delete = st.checkbox(
                f"⚠️ Hapus data {month} {year} dari sistem dan GitHub.",
                key=f"confirm_delete_{month}_{year}"
            )

            if st.button("🗑️ Hapus Data Ini", type="primary") and confirm_delete:
                try:
                    del st.session_state.data_storage[period_to_delete]
                    token = st.secrets.get("GITHUB_TOKEN")
                    repo_name = st.secrets.get("GITHUB_REPO")
                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)
                    contents = repo.get_contents(f"data/IKPA_{month}_{year}.xlsx")
                    repo.delete_file(contents.path, f"Delete {filename}", contents.sha)
                    st.success(f"✅ Data {month} {year} dihapus dari sistem & GitHub.")
                    st.snow()
                except Exception as e:
                    st.error(f"❌ Gagal menghapus data: {e}")

    # ============================================================
    # TAB 3: DOWNLOAD DATA
    # ============================================================
    with tab3:
        st.subheader("📥 Download Data IKPA")
        if not st.session_state.data_storage:
            st.info("ℹ️ Belum ada data.")
        else:
            available_periods = sorted(st.session_state.data_storage.keys(), reverse=True)
            period_to_download = st.selectbox(
                "Pilih periode untuk download",
                options=available_periods,
                format_func=lambda x: f"{x[0].capitalize()} {x[1]}"
            )
            df_download = st.session_state.data_storage[period_to_download]
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_excel = df_download.drop(['Bobot', 'Nilai Terbobot'], axis=1, errors='ignore')
                df_excel.to_excel(writer, index=False, sheet_name='Data IKPA')
            output.seek(0)
            st.download_button(
                label="📥 Download Excel",
                data=output,
                file_name=f"IKPA_{period_to_download[0]}_{period_to_download[1]}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # ============================================================
    # TAB 4: DOWNLOAD TEMPLATE (including Reference Template)
    # ============================================================
    with tab4:
        st.subheader("📋 Download Template")
        st.markdown("### 📘 Template IKPA")
        try:
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]
            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)
            file_content = repo.get_contents("templates/Template_IKPA.xlsx")
            template_data = base64.b64decode(file_content.content)
        except Exception:
            template_data = get_template_file()

        if template_data:
            st.download_button(
                label="📥 Download Template IKPA",
                data=template_data,
                file_name="Template_IKPA.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("---")
        st.markdown("### 📗 Template Data Referensi Satker & K/L")

        # 🧩 Use latest reference data for template content
        if 'reference_df' in st.session_state and not st.session_state.reference_df.empty:
            template_ref = st.session_state.reference_df.copy()
        else:
            # fallback: try load from GitHub
            try:
                token = st.secrets["GITHUB_TOKEN"]
                repo_name = st.secrets["GITHUB_REPO"]
                g = Github(auth=Auth.Token(token))
                repo = g.get_repo(repo_name)
                ref_content = repo.get_contents("templates/Template_Data_Referensi.xlsx")
                ref_data = base64.b64decode(ref_content.content)
                template_ref = pd.read_excel(io.BytesIO(ref_data))
            except Exception:
                template_ref = pd.DataFrame({
                    'No': [],
                    'Kode BA': [],
                    'K/L': [],
                    'Kode Satker': [],
                    'Uraian Satker-SINGKAT': [],
                    'Uraian Satker-LENGKAP': []
                })

        output_ref = io.BytesIO()
        with pd.ExcelWriter(output_ref, engine='openpyxl') as writer:
            template_ref.to_excel(writer, index=False, sheet_name='Data Referensi')
        output_ref.seek(0)

        st.download_button(
            label="📥 Download Template Data Referensi",
            data=output_ref,
            file_name="Template_Data_Referensi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # ============================================================
    # TAB 5: LOG AKTIVITAS
    # ============================================================
    with tab5:
        st.subheader("📖 Log Aktivitas GitHub")
        if not st.session_state.activity_log:
            st.info("Belum ada aktivitas.")
        else:
            df_log = pd.DataFrame(st.session_state.activity_log)
            st.dataframe(df_log[::-1].reset_index(drop=True), use_container_width=True)
            if st.button("🧹 Bersihkan Log"):
                st.session_state.activity_log = []
                st.success("🧹 Log dibersihkan.")


# ===============================
# 🔹 MAIN APP
# ===============================
def main():
    # ============================================================
    # 🧩 Auto-load Reference Data from GitHub FIRST
    # ============================================================
    if 'reference_df' not in st.session_state:
        try:
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]
            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)
            ref_path = "templates/Template_Data_Referensi.xlsx"
            ref_file = repo.get_contents(ref_path)
            ref_data = base64.b64decode(ref_file.content)
            ref_df = pd.read_excel(io.BytesIO(ref_data))
            ref_df['Kode Satker'] = ref_df['Kode Satker'].astype(str)
            st.session_state.reference_df = ref_df
            st.info(f"📚 Data Referensi dimuat otomatis ({len(ref_df)} baris).")
        except Exception as e:
            st.warning(f"⚠️ Tidak dapat memuat Data Referensi dari GitHub: {e}")

    # ============================================================
    # ✅ Then load data from GitHub (files can now be merged cleanly)
    # ============================================================
    if not st.session_state.get("data_storage"):
        with st.spinner("🔄 Memuat data dari GitHub..."):
            try:
                load_data_from_github()
            except Exception as e:
                st.error(f"⚠️ Gagal memuat data dari GitHub: {e}")

    # ===============================
    # 🔹 Sidebar Navigation
    # ===============================
    st.sidebar.title("🧭 Navigasi")
    st.sidebar.markdown("---")

    page = st.sidebar.radio(
        "Pilih Halaman",
        options=[
            "📊 Dashboard Utama",
            "📈 Dashboard Internal",
            "🔐 Admin"
        ],
        index=0
    )

    st.sidebar.markdown("---")
    st.sidebar.info("""
    **Dashboard IKPA**  
    Indikator Kinerja Pelaksanaan Anggaran  
    KPPN Baturaja

    📧 Support: ameer.noor@kemenkeu.go.id
    """)

    # ===============================
    # 🔹 Routing Halaman
    # ===============================
    if page == "📊 Dashboard Utama":
        try:
            page_dashboard()
        except Exception as e:
            st.error(f"❌ Terjadi kesalahan di Dashboard Utama: {e}")

    elif page == "📈 Dashboard Internal":
        try:
            page_trend()
        except Exception as e:
            st.error(f"❌ Terjadi kesalahan di Dashboard Internal KPPN: {e}")

    elif page == "🔐 Admin":
        try:
            page_admin()
        except Exception as e:
            st.error(f"❌ Terjadi kesalahan di Halaman Admin: {e}")

# ===============================
# 🔹 ENTRY POINT
# ===============================
if __name__ == "__main__":
    main()
